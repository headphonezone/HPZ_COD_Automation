import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill
from ui import apply_global_style, page_header

# -------------------------------------------------
# PAGE
# -------------------------------------------------

st.set_page_config(layout="wide",initial_sidebar_state="collapsed")
apply_global_style()
page_header("🚚 Delhivery COD Reconciliation")

# -------------------------------------------------
# SESSION STATE
# -------------------------------------------------

if "processed" not in st.session_state:
    st.session_state.processed = False

# -------------------------------------------------
# HELPERS
# -------------------------------------------------

def load_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    return pd.read_excel(file)

def clean_order(col):
    return (
        col.astype(str)
        .str.replace("#", "", regex=False)
        .str.replace(".0", "", regex=False)
        .str.replace("\xa0", "", regex=False)
        .str.strip()
    )

def clean_text(col):
    return (
        col.astype(str)
        .str.replace(".0", "", regex=False)
        .str.strip()
    )

# ---------- Excel Auto Width ----------
def dataframe_to_excel(df, sheet):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.book.active

        for column in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = length + 2

    buffer.seek(0)
    return buffer

# ---------- Lookup Highlight ----------
def create_lookup_excel(df):

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lookup")

        ws = writer.book.active

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        status_col = df.columns.get_loc("Status") + 1

        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=status_col).value
            fill = green if status == "Matched" else red

            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

    buffer.seek(0)
    return buffer

# -------------------------------------------------
# DATE PICKER
# -------------------------------------------------

st.subheader("Tally Posting Date")

posting_date = st.date_input(
    "Select Date for Tally Import",
    value=pd.to_datetime("today")
)

posting_date = posting_date.strftime("%d-%m-%Y")

# -------------------------------------------------
# FILE UPLOAD
# -------------------------------------------------

st.header("Upload Files")

col1, col2 = st.columns(2)

with col1:
    delhivery_file = st.file_uploader("Upload Delhivery File")

with col2:
    sales_file = st.file_uploader("Upload Sales File")

# -------------------------------------------------
# RUN RECONCILIATION
# -------------------------------------------------

if delhivery_file and sales_file:

    if st.button("🚀 Run Reconciliation"):

        progress = st.progress(0)
        status = st.empty()

        with st.spinner("Processing Files..."):

            # LOAD
            status.text("Loading files...")
            delhivery = load_file(delhivery_file)
            sales = load_file(sales_file)

            progress.progress(20)

            # CLEAN COLUMN NAMES
            delhivery.columns = delhivery.columns.str.strip()
            sales.columns = sales.columns.str.strip()

            # -------------------------------------------------
            # PRIMARY KEY CLEANING
            # -------------------------------------------------

            status.text("Preparing Order Numbers...")

            if "Order Number" not in delhivery.columns:
                st.error("❌ 'Order Number' column missing in Delhivery file")
                st.stop()

            if "Sale Order Number" not in sales.columns:
                st.error("❌ 'Sale Order Number' column missing in Sales file")
                st.stop()

            delhivery["ORDER_KEY"] = clean_order(delhivery["Order Number"])
            sales["ORDER_KEY"] = clean_order(sales["Sale Order Number"])

            progress.progress(40)

            # -------------------------------------------------
            # REMOVE DUPLICATES USING ORDER NUMBER
            # -------------------------------------------------

            sales_unique = sales.drop_duplicates("ORDER_KEY")

            # -------------------------------------------------
            # MERGE
            # -------------------------------------------------

            status.text("Reconciling records...")

            recon = delhivery.merge(
                sales_unique,
                on="ORDER_KEY",
                how="left",
                indicator=True
            )

            matched = recon[recon["_merge"] == "both"]
            unmatched = recon[recon["_merge"] == "left_only"]

            progress.progress(60)

            # -------------------------------------------------
            # TALLY IMPORT (MATCHED + UNMATCHED)
            # -------------------------------------------------

            status.text("Preparing Tally Import...")

            recon["Customer Email"] = recon["Customer Email"].fillna("N/A")

            tally_import = pd.DataFrame({
                "Date": posting_date,
                "Credit": recon["Customer Email"],
                "Debit": "Delhivery COD Receivable",
                "Debit Reference no": recon["ORDER_KEY"].fillna("N/A"),
                "Gross Total": recon["COD Amount"],
                "Narration": clean_text(recon["Waybill Number"])
            })

            tally_excel = dataframe_to_excel(
                tally_import,
                "Delhivery COD Import"
            )

            progress.progress(75)

            # -------------------------------------------------
            # UNMATCHED REPORT
            # -------------------------------------------------

            unmatched_report = unmatched[["ORDER_KEY"]].copy()
            unmatched_report["Issue"] = "Order not found in Sales"

            unmatched_excel = dataframe_to_excel(
                unmatched_report,
                "Unmatched"
            )

            progress.progress(85)

            # -------------------------------------------------
            # LOOKUP FILE
            # -------------------------------------------------

            status.text("Creating Lookup File...")

            lookup = recon.copy()

            lookup["Status"] = lookup["_merge"].map({
                "both": "Matched",
                "left_only": "Unmatched"
            })

            lookup.drop(columns="_merge", inplace=True)

            lookup_excel = create_lookup_excel(lookup)

            progress.progress(100)

        status.success("✅ Reconciliation Completed")

        st.session_state.processed = True
        st.session_state.tally_excel = tally_excel
        st.session_state.unmatched_excel = unmatched_excel
        st.session_state.lookup_excel = lookup_excel
        st.session_state.matched_count = len(matched)
        st.session_state.unmatched_count = len(unmatched)

# -------------------------------------------------
# RESULTS
# -------------------------------------------------

if st.session_state.processed:

    st.divider()
    st.success("🎉 Processing Completed")

    c1, c2 = st.columns(2)

    c1.metric("✅ Matched Records", st.session_state.matched_count)
    c2.metric("❌ Unmatched Records", st.session_state.unmatched_count)

    st.subheader("Download Outputs")

    d1, d2, d3 = st.columns(3)

    d1.download_button(
        "⬇ Delhivery COD Import",
        st.session_state.tally_excel,
        "Delhivery_COD_Import.xlsx"
    )

    d2.download_button(
        "⬇ Unmatched Report",
        st.session_state.unmatched_excel,
        "Unmatched_Report.xlsx"
    )

    d3.download_button(
        "⬇ Lookup File",
        st.session_state.lookup_excel,
        "Lookup_File.xlsx"
    )