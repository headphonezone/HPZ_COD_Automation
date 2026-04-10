import streamlit as st
import pandas as pd
import base64
from io import BytesIO
from openpyxl.styles import PatternFill
from ui import apply_global_style, page_header


# -------------------------------------------------
# PAGE
# -------------------------------------------------

st.set_page_config(layout="wide",initial_sidebar_state="collapsed")
apply_global_style()
page_header("📦 Bluedart COD Reconciliation")


# -------------------------------------------------
# SESSION STATE
# -------------------------------------------------

if "processed" not in st.session_state:
    st.session_state.processed = False

# -------------------------------------------------
# HELPERS
# -------------------------------------------------

def load_file(file):
    if file.name.endswith(".csv"):
        return pd.read_csv(file)
    return pd.read_excel(file)

def clean_awb(col):
    return (
        col.astype(str)
        .str.replace(".0", "", regex=False)
        .str.replace("\xa0", "", regex=False)
        .str.replace(r"\s+", "", regex=True)
        .str.strip()
    )

# ---------- Excel Creator ----------
def dataframe_to_excel(df, sheet="Sheet1"):
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet)

        ws = writer.book.active

        for column in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = length + 2

    buffer.seek(0)
    return buffer

# ---------- Lookup Highlight ----------
def create_lookup_excel(df):

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Lookup")

        ws = writer.book.active

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        status_col = df.columns.get_loc("Status") + 1

        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=status_col).value
            fill = green if status == "Matched" else red

            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

    buffer.seek(0)
    return buffer

# -------------------------------------------------
# FILE UPLOAD
# -------------------------------------------------

st.header("Upload Files")

col1, col2 = st.columns(2)

with col1:
    bluedart_file = st.file_uploader("Upload Bluedart File")

with col2:
    sales_file = st.file_uploader("Upload Sales File")

# -------------------------------------------------
# RUN RECONCILIATION
# -------------------------------------------------

if bluedart_file and sales_file:

    if st.button("🚀 Run Reconciliation"):

        progress = st.progress(0)
        status = st.empty()

        with st.spinner("Processing Files..."):

            # LOAD FILES
            status.text("Loading files...")
            sales = load_file(sales_file)

            bd = pd.read_excel(bluedart_file, header=1, skiprows=[2])
            bd.columns = bd.columns.str.strip()

            progress.progress(20)

            # CLEAN DATA
            status.text("Cleaning data...")

            awb_col = [c for c in bd.columns if "awb" in c.lower()][0]
            bd.rename(columns={awb_col: "AWB"}, inplace=True)

            bd.rename(columns={
                "AMOUNT": "COD Amount",
                "PAY REF DATE": "PAY REF DATE"
            }, inplace=True)

            bd["AWB"] = clean_awb(bd["AWB"])

            sales.rename(columns={"AWB num": "AWB"}, inplace=True)
            sales["AWB"] = clean_awb(sales["AWB"])

            progress.progress(40)

            # LEFT OUTER JOIN (CORRECT FINANCE LOGIC)
            status.text("Reconciling records...")

            sales_unique = sales.drop_duplicates("AWB")

            recon = bd.merge(
                sales_unique,
                on="AWB",
                how="left",
                indicator=True
            )

            matched = recon[recon["_merge"] == "both"]
            unmatched = recon[recon["_merge"] == "left_only"]

            progress.progress(60)

            # -------------------------------------------------
            # TALLY OUTPUT (MATCHED + UNMATCHED)
            # -------------------------------------------------

            status.text("Preparing Tally Upload...")

            tally = recon[
                ["AWB","Customer Email","Sale Order Number","COD Amount","PAY REF DATE"]
            ].copy()

            tally["Date"] = pd.to_datetime(
                tally["PAY REF DATE"], errors="coerce"
            ).dt.strftime("%d-%m-%Y")

            tally["Sale Order Number"] = (
                tally["Sale Order Number"]
                .astype(str)
                .str.replace("#","",regex=False)
            )

            # Replace missing values
            tally = tally.fillna("N/A")

            tally_import = pd.DataFrame({
                "Date": tally["Date"].replace("NaT", "N/A"),
                "Credit": tally["Customer Email"],
                "Debit": "COD Sales",
                "Debit Reference no": tally["Sale Order Number"],
                "Gross Total": tally["COD Amount"],
                "Narration": tally["AWB"]
            })

            tally_excel = dataframe_to_excel(tally_import, "Tally Upload")

            progress.progress(75)

            # UNMATCHED REPORT
            unmatched_report = unmatched[["AWB"]].copy()
            unmatched_report["Issue"] = "AWB not found in Sales"

            unmatched_excel = dataframe_to_excel(unmatched_report, "Unmatched")

            progress.progress(85)

            # LOOKUP FILE
            status.text("Creating Lookup File...")

            lookup = recon.copy()

            lookup["Status"] = lookup["_merge"].map({
                "both": "Matched",
                "left_only": "Unmatched"
            })

            lookup.drop(columns="_merge", inplace=True)

            lookup_excel = create_lookup_excel(lookup)

            progress.progress(100)

        status.success("✅ Reconciliation Completed")

        # SAVE SESSION STATE
        st.session_state.processed = True
        st.session_state.tally_excel = tally_excel
        st.session_state.unmatched_excel = unmatched_excel
        st.session_state.lookup_excel = lookup_excel
        st.session_state.matched_count = len(matched)
        st.session_state.unmatched_count = len(unmatched)

# -------------------------------------------------
# RESULTS + COUNTS + DOWNLOADS
# -------------------------------------------------

if st.session_state.processed:

    st.divider()
    st.success("🎉 Processing Completed")

    c1, c2 = st.columns(2)

    with c1:
        st.metric("✅ Matched Records", st.session_state.matched_count)

    with c2:
        st.metric("❌ Unmatched Records", st.session_state.unmatched_count)

    st.subheader("Download Outputs")

    d1, d2, d3 = st.columns(3)

    with d1:
        st.download_button(
            "⬇ Tally Upload (Excel)",
            data=st.session_state.tally_excel,
            file_name="Tally_Upload.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with d2:
        st.download_button(
            "⬇ Unmatched Report",
            data=st.session_state.unmatched_excel,
            file_name="Unmatched_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with d3:
        st.download_button(
            "⬇ Lookup File",
            data=st.session_state.lookup_excel,
            file_name="Lookup_File.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )